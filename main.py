# Modules
import logging
import sys
import os
from urllib.parse import urlparse, unquote
import pandas as pd
import sharepy
from openpyxl import load_workbook
from io import StringIO
import boto3
from botocore.exceptions import ClientError

logger = logging.getLogger()
logger.setLevel(logging.INFO)
logging.getLogger().addHandler(logging.StreamHandler(sys.stdout))


def handler(event, context):
    # Lambda Entry Point
    logger.info('## ENVIRONMENT VARIABLES')
    logger.info(os.environ)
    logger.info('## EVENT')
    logger.info(event)

    tablesEnv = [i.strip() for i in list(filter(None, (os.environ['tables'].split(","))))]
    worksheetsEnv = [i.strip() for i in list(filter(None, (os.environ['worksheets'].split(","))))]

    excel = downloadExcel(os.environ['excelUrl'])
    parseExcel(excel, table=tablesEnv, worksheet=worksheetsEnv)


def downloadExcel(url):
    logger.info('url: {0}'.format(url))

    urlParse = urlparse(url)
    fileName = unquote(os.path.basename(urlParse.path)).replace(" ", "_")

    ssm = boto3.client('ssm')
    try:
        user = ssm.get_parameter(Name=os.environ['userSSMParam'], WithDecryption=True)
        passwd = ssm.get_parameter(Name=os.environ['passwdSSMParam'], WithDecryption=True)
    except ClientError as e:
        logger.critical("Unexpected error: {0}".format(e))
        raise Exception()

    try:
        sharePy = sharepy.connect(urlParse.netloc, username=user['Parameter']['Value'],
                                  password=passwd['Parameter']['Value'])
    except:
        e = sys.exc_info()[0]
        logger.critical("Unexpected error connecting to Sharepoint: {0}".format(e))
        raise Exception()

    try:
        sharePy.getfile(url, filename="/tmp/" + fileName)
    except:
        e = sys.exc_info()[0]
        logger.critical("Unexpected error downloading {0} from Sharepoint: {1}".format(fileName, e))
        raise Exception()

    return fileName


def parseExcel(file: str, worksheet: list = None, table: list = None):
    wb = load_workbook("/tmp/" + file, data_only=True)

    if not worksheet:
        logger.info("worksheets found: {0}".format(wb.sheetnames))
        sheets = wb.sheetnames
    else:
        logger.info("worksheets defined by user: {0}".format(worksheet))
        sheets = worksheet

    for sheet in sheets:

        ws = wb[sheet]
        if ws.tables.items():
            wsTable = []

            if not table:
                logger.info(
                    "Tables found: {0} on sheet: {1}".format({key: value for key, value in ws.tables.items()},
                                                             ws.title))
                wsTable = ws.tables.items()
            elif len(table) > 0:
                if any(item in table for item in ws.tables.keys()):
                    for m in table:
                        for n in ws.tables.keys():
                            if m == n:
                                logger.info("Table defined by user found: {0}".format(m))
                                wsTable.append((ws.tables[m].name, ws.tables[m].ref))
                else:
                    logger.info("User defined table '{0}' not found on this worksheet: {1}".format(table, sheet))
                    wsTable = []
            else:
                e = sys.exc_info()[0]
                logger.critical("Here be dragons. {0}".format(e))
                raise Exception()

            for entry, data_boundary in wsTable:

                data = ws[data_boundary]
                content = []
                for ent in data:
                    content.append([cell.value for cell in ent])

                header = content[0]
                rest = content[1:]

                df = pd.DataFrame(rest, columns=header)

                csv_buffer = StringIO()
                df.to_csv(csv_buffer, index=False)

                s3_copy = boto3.resource('s3')
                s3_copy.Object(os.environ['bucketName'], file + "_" + ws.title.replace(" ", "_") + ".csv").put(
                    Body=csv_buffer.getvalue())
        else:
            logger.info("no tables found on sheet {0}".format(ws))


def main():
    # main() is useful for testing locally. not used when run in Lambda
    test_event = {
        "version": "0",
        "account": "123456789012",
        "region": "eu-west-1",
        "detail": {},
        "detail-type": "Scheduled Event",
        "source": "aws.events",
        "time": "2019-03-01T01:23:45Z",
        "id": "cdc73f9d-aea9-11e3-9d5a-835b769c0d9c",
        "resources": [
            "arn:aws:events:eu-west-1:123456789012:rule/my-schedule"
        ]
    }
    logger.info(handler(test_event, ""))


if __name__ == '__main__':
    # Local Runtime entry point.
    main()
